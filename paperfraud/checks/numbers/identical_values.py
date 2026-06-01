"""Identical value detection across different experimental groups.

A common fabrication pattern: copying numeric values between supposedly
independent groups/analytes. Extracts numbers from Results text, clusters
them by nearby context, and flags repeated values with high precision.
"""

from __future__ import annotations

import re
from collections import defaultdict

from paperfraud.base import CheckResult, SourceLocation


def _norm(v: float) -> str:
    """Normalize a float to its full string representation for exact matching."""
    return f"{v:.10f}".rstrip('0').rstrip('.')


def extract_values_with_context(text: str) -> list[dict]:
    """Extract numeric values with surrounding context window.

    Returns list of {value, context, start_pos} dicts.
    """
    pattern = re.compile(r'(?<!\d)(\d+\.\d+)(?!\d)')
    results = []
    for match in pattern.finditer(text):
        val_str = match.group(0)
        if re.match(r'^(19|20)\d{2}$', val_str):
            continue
        if re.match(r'^\d+\.\d+(?:\.\d+)+$', val_str):
            continue
        try:
            val = float(val_str)
        except ValueError:
            continue
        start = max(0, match.start() - 80)
        end = min(len(text), match.end() + 80)
        context = text[start:end].replace('\n', ' ').strip()
        results.append({
            "value": val,
            "value_str": val_str,
            "context": context,
            "start_pos": match.start(),
        })
    return results


# P-value thresholds and common statistical cutoffs — these appear everywhere
_PVALUE_THRESHOLDS = {
    0.05, 0.01, 0.001, 0.0001,
    0.5, 1.0, 1.96,
}
# Numbers that are clearly P-value thresholds when near "P <" or "P >"
_PVALUE_CONTEXT_RE = __import__('re').compile(
    r'P\s*[<>≤≥]\s*\d+\.\d+', __import__('re').IGNORECASE
)

# DOI prefix pattern: 10.XXXX... (4+ digits after the dot)
_DOI_PATTERN = __import__('re').compile(r'\b10\.\d{4,}')

# Virus variant/lineage pattern: letter.digits (e.g., B.1.351, BA.1.1, P.1)
_VARIANT_PATTERN = __import__('re').compile(
    r'\b[A-Z]{1,4}\.\d+(?:\.\d+)*\b'
)


def _is_pvalue_threshold(v: float, ctx: str) -> bool:
    """Check if this value is a P-value threshold rather than experimental data."""
    for t in _PVALUE_THRESHOLDS:
        if abs(v - t) < 1e-9:
            return True
    if _PVALUE_CONTEXT_RE.search(ctx):
        return True
    return False


def _is_metadata_number(v_str: str, ctx: str) -> bool:
    """Check if a number is metadata (DOI, variant name) rather than data."""
    if _DOI_PATTERN.search(ctx):
        return True
    if _VARIANT_PATTERN.search(ctx):
        return True
    return False


def _float_to_key(v: float, precision: int) -> str:
    """Convert float to string key at given precision for dedup."""
    return f"{v:.{precision}f}"


def detect_identical_values(values_with_ctx: list[dict]) -> dict:
    """Find identical values appearing in different contextual windows.

    Filters out:
      1. P-value thresholds (0.05, 0.01, etc.)
      2. Values clustered in the same table (all occurrences within 4000 chars
         or context windows sharing >50% word overlap — typical of large
         comparison matrices like genome identity tables)

    Returns dict with:
      - duplicates: list of {value, precision, occurrences, is_table}
      - suspicious_count: number of values repeated >= 3 times (excluding table clusters)
      - table_clustered: number of duplicate groups that are same-table artifacts
    """
    # Filter out P-value thresholds and metadata numbers (DOIs, variant names)
    filtered = [
        e for e in values_with_ctx
        if not _is_pvalue_threshold(e["value"], e["context"])
        and not _is_metadata_number(e["value_str"], e["context"])
    ]

    by_precision: dict[int, dict[str, list[dict]]] = defaultdict(lambda: defaultdict(list))

    for entry in filtered:
        v = entry["value"]
        for prec in [2, 3, 4]:
            key = _float_to_key(v, prec)
            by_precision[prec][key].append(entry)

    def _is_table_cluster(entries: list[dict]) -> bool:
        """Check if all duplicates come from the same table/block."""
        if len(entries) < 2:
            return False
        positions = [e["start_pos"] for e in entries]
        char_span = max(positions) - min(positions)
        # All within 4000 chars → likely same table
        if char_span < 4000:
            return True
        # Check word overlap between context windows
        if len(entries) >= 3:
            word_sets = []
            for e in entries:
                words = set(re.sub(r'[^a-zA-Z]', ' ', e["context"].lower()).split())
                word_sets.append(words)
            # If all pairs share >50% words, it's the same table
            overlaps = 0
            total_pairs = 0
            for i in range(len(word_sets)):
                for j in range(i + 1, len(word_sets)):
                    if not word_sets[i] or not word_sets[j]:
                        continue
                    total_pairs += 1
                    intersection = word_sets[i] & word_sets[j]
                    union = word_sets[i] | word_sets[j]
                    if len(intersection) / max(len(union), 1) > 0.5:
                        overlaps += 1
            if total_pairs > 0 and overlaps / total_pairs > 0.6:
                return True
        return False

    # Find duplicates
    duplicates = []
    for prec in [2, 3, 4]:
        for key, entries in by_precision[prec].items():
            if len(entries) < 2:
                continue
            contexts = [e["context"] for e in entries]
            unique_contexts = set()
            for ctx in contexts:
                norm_ctx = re.sub(r'[^a-zA-Z]', ' ', ctx.lower())
                norm_ctx = ' '.join(norm_ctx.split())
                unique_contexts.add(norm_ctx[:60])
            if len(unique_contexts) >= 2:
                is_table = _is_table_cluster(entries)
                duplicates.append({
                    "value": float(key),
                    "precision": prec,
                    "count": len(entries),
                    "contexts": contexts[:5],
                    "is_table": is_table,
                })

    # Deduplicate: if same value flagged at multiple precisions, keep highest
    seen_vals: set[float] = set()
    unique_dups = []
    for d in sorted(duplicates, key=lambda x: x["precision"], reverse=True):
        if d["value"] not in seen_vals:
            unique_dups.append(d)
            seen_vals.add(d["value"])

    # Separate table artifacts from real signals
    table_dups = [d for d in unique_dups if d["is_table"]]
    real_dups = [d for d in unique_dups if not d["is_table"]]

    pvalue_hits = sum(1 for e in values_with_ctx if _is_pvalue_threshold(e["value"], e["context"]))
    metadata_hits = sum(1 for e in values_with_ctx if _is_metadata_number(e["value_str"], e["context"]))

    return {
        "duplicates": real_dups,  # Only return non-table duplicates as main result
        "table_duplicates": table_dups,
        "suspicious_count": sum(1 for d in real_dups if d["count"] >= 3),
        "total_duplicate_values": len(real_dups),
        "total_values_checked": len(filtered),
        "pvalue_filtered": pvalue_hits,
        "metadata_filtered": metadata_hits,
        "table_filtered": len(table_dups),
    }


def _col_letter(idx: int) -> str:
    """Convert 1-indexed column number to Excel letter(s)."""
    s = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        s = chr(65 + rem) + s
    return s


def _is_sequence(values: list[float], tolerance: float = 0.05) -> bool:
    """Check if values form a geometric or arithmetic progression.

    Used to identify parameter columns (dose/concentration/time series)
    so their values aren't flagged as duplicate fraud.
    """
    if len(values) < 3:
        return False
    uniq = sorted(set(values))
    if len(uniq) < 3:
        return False

    # Try geometric progression (each value ≈ prev * ratio)
    ratios = []
    for i in range(len(uniq) - 1):
        if uniq[i] == 0:
            break
        ratios.append(uniq[i + 1] / uniq[i])
    if len(ratios) >= 2:
        mean = sum(ratios) / len(ratios)
        if all(abs(r - mean) / max(mean, 1e-10) < tolerance for r in ratios):
            return True

    # Try arithmetic progression (each value ≈ prev + diff)
    diffs = [uniq[i + 1] - uniq[i] for i in range(len(uniq) - 1)]
    if len(diffs) >= 2:
        mean = sum(diffs) / len(diffs)
        if abs(mean) < 1e-10:
            if all(abs(d) < 1e-10 for d in diffs):
                return True
        elif all(abs(d - mean) / abs(mean) < tolerance for d in diffs):
            return True

    return False


def detect_identical_values_in_table(data_file: str) -> dict:
    """Scan Excel/CSV source data for cell-level duplicate numeric values.

    Finds high-precision values that repeat across different columns
    (i.e. supposedly independent experimental groups). Values that only
    repeat within the same column (legitimate dose/parameter reuse) are
    excluded.

    Returns dict with:
      - duplicates: list of {value, precision, count, sheet, positions}
      - total_checked: total unique numeric values scanned
      - sheets_scanned: list of sheet names
    """
    from pathlib import Path

    path = Path(data_file)
    suffix = path.suffix.lower()

    if suffix not in (".xlsx", ".xls", ".csv", ".tsv"):
        return {"duplicates": [], "total_checked": 0, "sheets_scanned": [], "error": f"不支持的文件格式: {suffix}"}

    # ── Read data ──────────────────────────────────────────────────────────
    sheets: dict[str, list[list]] = {}  # sheet_name → 2D grid of cell values

    if suffix in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {"duplicates": [], "total_checked": 0, "sheets_scanned": [], "error": "需要安装 openpyxl: pip install openpyxl"}

        wb = load_workbook(path, data_only=True)
        for name in wb.sheetnames:
            ws = wb[name]
            grid = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
                grid.append(list(row))
            sheets[name] = grid
        wb.close()
    else:
        import csv
        rows = []
        with open(path, encoding="utf-8-sig") as f:
            reader = csv.reader(f, delimiter="\t" if suffix == ".tsv" else ",")
            for row in reader:
                rows.append(row)
        sheets[path.stem] = rows

    # ── Collect numeric values with positions ───────────────────────────────
    Position = tuple[str, int, int, float]  # (sheet, row, col, raw_value)

    from collections import defaultdict
    value_positions: dict[str, list[Position]] = defaultdict(list)
    total_checked = 0

    for sheet_name, grid in sheets.items():
        for r, row in enumerate(grid):
            if row is None:
                continue
            for c, cell in enumerate(row):
                if cell is None:
                    continue
                if not isinstance(cell, (int, float)):
                    continue
                v = float(cell)
                # Normalize to full-precision string key
                key = f"{v:.15f}".rstrip("0")
                value_positions[key].append((sheet_name, r, c, v))
                total_checked += 1

    # ── Identify parameter columns (dose/concentration/time series) ────────
    # Round to 4dp for sequence detection: Excel often stores the same intended
    # value (e.g. 100/3) with different float precision across blocks
    # (33.33333 vs 33.333333 vs 33.3333333333333), which would break
    # geometric progression detection and create phantom duplicates.
    col_values: dict[tuple[str, int], list[float]] = defaultdict(list)
    row_values: dict[tuple[str, int], list[float]] = defaultdict(list)
    for (_key, positions) in value_positions.items():
        for sheet, row, col, val in positions:
            rounded = round(val, 4)
            col_values[(sheet, col)].append(rounded)
            row_values[(sheet, row)].append(rounded)

    param_columns: set[tuple[str, int]] = set()
    for (sheet, col), vals in col_values.items():
        if _is_sequence(vals):
            param_columns.add((sheet, col))

    # Also detect horizontal parameter rows (e.g. dose series across columns)
    param_rows: set[tuple[str, int]] = set()
    for (sheet, row), vals in row_values.items():
        if _is_sequence(vals):
            param_rows.add((sheet, row))

    # ── Detect template-level repetition in large structured sheets ──────
    # In omics datasets, the same value often repeats N times because of N
    # condition/analysis columns — each column block reuses the same gene
    # list or reference values. If most duplicates from a sheet share the
    # same repeat count, that count is template structure, not fabrication.
    sheet_count_dist: dict[str, dict[int, int]] = defaultdict(lambda: defaultdict(int))
    for key, positions in value_positions.items():
        if len(positions) < 2:
            continue
        for sheet, _, _, _ in positions:
            sheet_count_dist[sheet][len(positions)] += 1

    template_counts: set[int] = set()
    for sheet, dist in sheet_count_dist.items():
        if not dist:
            continue
        total = sum(dist.values())
        dominant_count, dominant_freq = max(dist.items(), key=lambda x: x[1])
        # If a single repeat count dominates (>60% of duplicates in this sheet
        # and >100 instances), treat it as template structure
        if dominant_freq > total * 0.6 and dominant_freq > 100:
            template_counts.add(dominant_count)

    # ── Find duplicates across different columns ────────────────────────────
    duplicates = []

    for key, positions in value_positions.items():
        if len(positions) < 2:
            continue

        # Skip extremely small values: statistical outputs (p-values, q-values)
        # that can legitimately repeat across different comparisons
        v = positions[0][3]
        if abs(v) < 1e-10 and v != 0:
            continue

        # Skip template-structure repetition (omics datasets)
        if len(positions) in template_counts:
            continue

        # Skip values that repeat too many times (> 8 distinct rows):
        # these are reused parameters (dose series, time points), not data fabrication
        distinct_rows = len({p[1] for p in positions})
        if distinct_rows > 8:
            continue

        # Check if all positions are in the same column → skip (legitimate reuse)
        columns = {p[2] for p in positions}
        if len(columns) == 1:
            continue

        # Skip if ANY position is in a parameter column/row — the value is a
        # legitimate dose/concentration/time parameter that may also appear
        # in other sheets/columns (e.g. Supplementary tables)
        pos_cols = [(p[0], p[2]) for p in positions]
        pos_rows = [(p[0], p[1]) for p in positions]
        if any(pc in param_columns for pc in pos_cols):
            continue
        if any(pr in param_rows for pr in pos_rows):
            continue

        # Check precision: count decimal places, cap at 8dp to avoid
        # float representation artifacts (8.14 → 8.140000000000001).
        # Real biological data beyond 8dp is vanishingly rare.
        v_str = f"{v:.8f}".rstrip("0").rstrip(".")
        if "." in v_str:
            decimals = len(v_str.split(".")[1])
        else:
            decimals = 0

        # Group by sheet for display
        sheets_with_dups = defaultdict(list)
        for p in positions:
            sheets_with_dups[p[0]].append(p)

        duplicates.append({
            "value": v,
            "precision": min(decimals, 10),
            "count": len(positions),
            "sheets": sheets_with_dups,
            "positions": [(p[0], p[1] + 1, _col_letter(p[2] + 1)) for p in positions],
        })

    # Sort: precision × count (higher is more suspicious)
    duplicates.sort(key=lambda d: (d["precision"] * d["count"]), reverse=True)

    return {
        "duplicates": duplicates,
        "total_checked": total_checked,
        "sheets_scanned": list(sheets.keys()),
    }


def run_identical_values(paper) -> list[CheckResult]:
    """Detect identical numeric values across different experimental contexts."""
    MAX_TEXT = 500_000
    text = paper.results or paper.full_text
    text_truncated = len(text) > MAX_TEXT
    if text_truncated:
        text = text[:MAX_TEXT]

    if not text:
        return [
            CheckResult(
                check_id="numbers.identical_values",
                check_name="跨组相同数值检测",
                level="error",
                verdict="无法执行：未提取到文本",
                needs_human=False,
            )
        ]

    values_with_ctx = extract_values_with_context(text)
    if len(values_with_ctx) < 10:
        result = CheckResult(
            check_id="numbers.identical_values",
            check_name="跨组相同数值检测",
            level="error",
            verdict=f"数值样本不足（仅提取到 {len(values_with_ctx)} 个小数，需 ≥ 10）",
            needs_human=False,
        )
        return [result]

    detection = detect_identical_values(values_with_ctx)

    evidence = []
    for d in detection["duplicates"][:10]:
        evidence.append(
            f"数值 {d['value']} 在 {d['count']} 处不同上下文中出现（精度: {d['precision']} 位小数）"
        )
        for ctx in d["contexts"][:3]:
            evidence.append(f"  上下文: ...{ctx[:120]}...")

    trunc_note = "（文本已截断）" if text_truncated else ""
    total_dup = detection["total_duplicate_values"]
    suspicious = detection["suspicious_count"]

    pval_info = f"（已过滤 {detection['pvalue_filtered']} 个 P 值/统计阈值"
    if detection.get("metadata_filtered", 0) > 0:
        pval_info += f"，{detection['metadata_filtered']} 个 DOI/编号"
    if detection.get("table_filtered", 0) > 0:
        pval_info += f"，{detection['table_filtered']} 组表格重复"
    pval_info += "）"

    # ── Excel/CSV source data scan ──────────────────────────────────────────
    table_evidence: list[str] = []
    table_suspicious = 0
    data_file = getattr(paper, "data_file", "") or ""

    if data_file and data_file.lower().endswith((".xlsx", ".xls", ".csv", ".tsv")):
        table_detection = detect_identical_values_in_table(data_file)
        if table_detection.get("error"):
            table_evidence.append(f"表格扫描失败: {table_detection['error']}")
        else:
            sheet_list = ", ".join(table_detection["sheets_scanned"])
            table_evidence.append(
                f"已扫描 {len(table_detection['sheets_scanned'])} 个数据表 "
                f"（{sheet_list}），共 {table_detection['total_checked']} 个单元格"
            )

            # Show high-precision (>= 4dp) duplicates with count >= 3.
            # count=2 pairs are too weak — a single repetition could be
            # coincidental or template-level duplication.
            high_prec = [d for d in table_detection["duplicates"] if d["precision"] >= 4]
            strong_signals = [d for d in high_prec if d["count"] >= 3]
            table_suspicious = len(strong_signals)

            for d in strong_signals[:8]:
                pos_list = [f"{p[2]}{p[1]}" for p in d["positions"][:6]]
                pos_str = ", ".join(pos_list)
                if len(d["positions"]) > 6:
                    pos_str += f" 等 {len(d['positions'])} 处"
                table_evidence.append(
                    f"数值 {d['value']}（{d['precision']} 位精度）在 {d['count']} 个非同行/同列单元格重复 → {pos_str}"
                )

    # ── Merge text + table results ──────────────────────────────────────────
    evidence.extend(table_evidence)
    total_suspicious = suspicious + table_suspicious

    if total_suspicious >= 5:
        level = "red"
    elif total_suspicious >= 3:
        level = "red" if total_dup > 0 else "orange"
    elif total_dup >= 2 or table_suspicious >= 2:
        level = "orange"
    elif total_dup == 1 or table_suspicious == 1:
        level = "yellow"
    else:
        level = "green"

    if total_suspicious >= 3:
        verdict = f"发现 {total_suspicious} 组高精度数值重复（{suspicious} 处文本 + {table_suspicious} 处表格）{pval_info}{trunc_note}"
    elif total_dup >= 2:
        verdict = f"发现 {total_dup} 组数值在不同上下文中重复出现 {pval_info}{trunc_note}"
    elif total_dup == 1:
        verdict = f"发现 1 组数值在不同上下文中重复 {pval_info}{trunc_note}"
    elif table_suspicious >= 1:
        verdict = f"表格中发现 {table_suspicious} 处高精度数值重复，文本中未发现异常{pval_info}{trunc_note}"
    else:
        verdict = (
            f"未发现跨组重复数值（共检测 {detection['total_values_checked']} 个非阈值数值"
            f"，过滤 {detection['pvalue_filtered']} 个 P 值/统计阈值）{trunc_note}"
        )

    return [
        CheckResult(
            check_id="numbers.identical_values",
            check_name="跨组相同数值检测",
            level=level,
            verdict=verdict,
            evidence=evidence[:20],
            confidence=0.85,
            needs_human=(total_suspicious > 0),
            human_instruction="核对重复数值的原始上下文。若不同分析物/实验组出现完全相同数值（精度 ≥ 2 位小数），"
            "需向作者索取原始数据确认。注意区分合理重复（如相同浓度、时间点）。"
            "表格来源数据若在 ≥4 位精度下跨列重复，极可能为数据复制粘贴。",
        )
    ]
